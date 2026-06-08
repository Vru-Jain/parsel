"""
Integration tests against the ACTUAL customer demo documents.

These guard the exact files we will demo with:
  * Book 1.pdf            digital MAN B&W manual (must extract clean rows)
  * 05 1.pdf              scanned image-only manual (must be detected, not silent)
  * New Format (1) 2.xlsx target schema (output must match it)

Each test skips cleanly if the demo folder is absent, so CI without the files
still passes. Heavy tests are marked `slow`.
"""
from __future__ import annotations

import os

import pytest

from engine.pdf_extractor import extract_pdf
from engine.semantic_mapper import SemanticMapper
from engine.transformer import Transformer
from tests.conftest import DEMO_FILES, requires_demo


@requires_demo
class TestScannedManual:
    def test_05_detected_as_scanned(self, cfg):
        path = DEMO_FILES["manual_scanned"]
        if not os.path.exists(path):
            pytest.skip("05 1.pdf not present")
        cfg["options"]["enable_ocr"] = False
        res = extract_pdf(path, cfg)
        # it's an image-only manual: most/all pages have no text layer
        assert res.is_scanned
        assert len(res.scanned_pages) >= res.total_pages * 0.6
        # and we did NOT silently produce a fake-empty success
        assert res.combined.empty


@requires_demo
class TestDigitalManual:
    @pytest.fixture(scope="class")
    def extraction(self, request):
        import json
        cfgpath = os.path.join(os.path.dirname(os.path.dirname(__file__)), "config.json")
        cfg = json.load(open(cfgpath, encoding="utf-8"))
        path = DEMO_FILES["manual_digital"]
        if not os.path.exists(path):
            pytest.skip("Book 1.pdf not present")
        # Extract a representative section (not all 492 pages — high-quality
        # table detection is ~0.5s/page). This section contains real parts tables.
        return extract_pdf(path, cfg, page_range=(90, 130))

    @pytest.mark.slow
    def test_extracts_pages_in_range(self, extraction):
        assert len(extraction.frames) > 20  # most pages in the range have tables

    @pytest.mark.slow
    def test_headers_are_clean_not_drawing_numbers(self, extraction):
        # No frame should have a drawing-number-like header dominating —
        # the scorer should have rejected those in favor of real headers.
        bad = 0
        for fr in extraction.frames:
            cols = [str(c) for c in fr.dataframe.columns
                    if not str(c).startswith("__")]
            for c in cols:
                cl = c.strip().lower()
                if cl.replace("-", "").isdigit() and len(cl) >= 6:
                    bad += 1
        # allow a tiny tail of odd pages
        assert bad <= len(extraction.frames) * 0.05, f"{bad} drawing-number headers"

    @pytest.mark.slow
    def test_known_columns_appear(self, extraction, cfg):
        """A part-name-like and a position-like header should be detectable,
        whichever extraction pass won (find_tables -> 'Item Designation',
        fast text-block pass -> 'Description'/'No.'). Both map to the schema."""
        from engine.semantic_mapper import SemanticMapper
        mapper = SemanticMapper(cfg, "models")
        seen = set()
        for fr in extraction.frames:
            for c in fr.dataframe.columns:
                if not str(c).startswith("__"):
                    seen.add(str(c))
        mapping = mapper.map_columns(list(seen)).mapping
        targets = set(mapping.values())
        assert "Part Name" in targets, f"no Part Name header among {sorted(seen)}"
        assert "DrawingPosNo" in targets, f"no position header among {sorted(seen)}"

    @pytest.mark.slow
    def test_end_to_end_rows_have_part_names(self, extraction, cfg):
        mapper = SemanticMapper(cfg, "models")
        tr = Transformer(cfg)
        raw = extraction.combined
        # transform a sample slice for speed
        sample = raw.head(500)
        mr = mapper.map_columns(list(sample.columns))
        out = tr.transform(sample, mr.mapping)
        non_empty_names = out["Part Name"].astype(str).str.strip().ne("").sum()
        assert non_empty_names > 50

    @pytest.mark.slow
    def test_plate_numbers_and_titles_populated(self, extraction, cfg):
        """DrawingNo (plate number) filled on most rows, and Spare Group is a
        real section title — never a part description like 'Screw'/'Cover'."""
        mapper = SemanticMapper(cfg, "models")
        tr = Transformer(cfg)
        out = tr.transform(extraction.combined, mapper.map_columns(
            list(extraction.combined.columns)).mapping)
        n = len(out)
        assert n > 50
        drawing_filled = out["DrawingNo"].astype(str).str.strip().ne("").mean()
        assert drawing_filled > 0.8, f"DrawingNo only {drawing_filled:.0%} filled"
        # Spare Group must never be a bare part-ish word (title leak guard)
        groups = set(out["Spare Group"].astype(str).str.strip())
        for leak in ("Screw", "Cover", "Nut", "Gasket", "Drain pipe"):
            assert leak not in groups, f"part name '{leak}' leaked into Spare Group"


@requires_demo
class TestTargetSchema:
    def test_schema_matches_exactly(self, cfg):
        import openpyxl
        target = DEMO_FILES["target_xlsx"]
        if not os.path.exists(target):
            pytest.skip("target xlsx not present")
        wb = openpyxl.load_workbook(target, data_only=True)
        ws = wb["Sheet1"]
        headers = [ws.cell(row=1, column=c).value
                   for c in range(1, ws.max_column + 1)]
        assert cfg["target_schema"] == headers
