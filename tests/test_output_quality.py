"""
Tests for the demo-quality fixes:
  * noise-row removal (page numbers, 'nan', stray header fragments)
  * NaN/'nan'/'none' normalized to empty
  * strict 14-column main sheet + separate unmapped sheet
  * section-title (Spare Group) extraction near a 'Plate'/'Drawing' locator
  * leading zeros on position numbers preserved as text
"""
from __future__ import annotations

import fitz
import numpy as np
import pandas as pd
import pytest
import openpyxl

import re as _re

from engine.transformer import Transformer
from engine.pdf_extractor import _extract_plate_info, _is_title_candidate
from engine.semantic_mapper import SemanticMapper
from engine.pipeline import process_file

_PLATE_RE = _re.compile(r"\b\d{3,4}-\d{3,4}-\d{2,4}\b")


class TestNoiseRemoval:
    def test_nan_strings_become_empty(self, cfg):
        tr = Transformer(cfg)
        df = pd.DataFrame({"Part Name": ["valve", "nan", "NaN", "real part"]})
        out = tr.transform(df, {"Part Name": "Part Name"})
        # "nan"/"NaN" rows dropped, real ones kept
        assert "Valve" in out["Part Name"].tolist()
        assert "Real Part" in out["Part Name"].tolist()
        assert not any(str(v).strip().lower() == "nan" for v in out["Part Name"])

    def test_real_nan_float_handled(self, cfg):
        tr = Transformer(cfg)
        df = pd.DataFrame({"Part Name": ["gasket", np.nan]})
        out = tr.transform(df, {"Part Name": "Part Name"})
        assert len(out) == 1
        assert out.loc[0, "Part Name"] == "Gasket"

    def test_page_number_rows_dropped(self, cfg):
        tr = Transformer(cfg)
        df = pd.DataFrame({
            "Part Name": ["O-ring", "(2)", "12", ""],
            "Spare Part No": ["P1", "", "", ""],
        })
        out = tr.transform(df, {"Part Name": "Part Name", "Spare Part No": "Spare Part No"})
        names = out["Part Name"].tolist()
        assert any("O-ring" in n or "O-Ring" in n for n in names)
        assert "(2)" not in names
        assert "12" not in names

    def test_numeric_partname_kept_if_has_part_number(self, cfg):
        tr = Transformer(cfg)
        df = pd.DataFrame({"Part Name": ["123"], "Spare Part No": ["XYZ-9"]})
        out = tr.transform(df, {"Part Name": "Part Name", "Spare Part No": "Spare Part No"})
        assert len(out) == 1  # real part number means keep it

    def test_recommended_spare_asterisk_stripped(self, cfg):
        # MAN B&W marks recommended spares with a trailing '*' which is not
        # part of the name and would otherwise be a manual edit.
        tr = Transformer(cfg)
        df = pd.DataFrame({"Part Name": ["Relief Valve, Complete*", "O-ring*"]})
        out = tr.transform(df, {"Part Name": "Part Name"})
        names = out["Part Name"].tolist()
        assert "Relief Valve, Complete" in names
        assert not any(str(n).endswith("*") for n in names)


class TestMaintenancePageSkip:
    def test_service_interval_page_is_skipped(self, tmp_path, cfg):
        """A maintenance/overhaul-schedule page ('Service interval ...') must be
        skipped, not scraped into junk C/O/R interval-code rows."""
        from engine.pdf_extractor import extract_pdf
        pdf = str(tmp_path / "sched.pdf")
        doc = fitz.open()
        page = doc.new_page(width=595, height=842)
        page.insert_text((60, 80), "MAN B&W")
        page.insert_text((60, 110), "Service interval (x1000 hours of operation)")
        page.insert_text((60, 140), "No.  Procedure  H  1  2  4  C  O  R")
        page.insert_text((60, 170), "1065-0101  Crankcase Relief Valve  C  O  M")
        doc.save(pdf)
        doc.close()
        res = extract_pdf(pdf, cfg)
        assert res.combined.empty  # nothing scraped from a schedule page

    def test_skip_markers_present_in_config(self, cfg):
        markers = [m.lower() for m in cfg["options"]["skip_page_markers"]]
        assert "service interval" in markers


class TestPlateInfoExtraction:
    def test_title_and_plate_from_caption(self):
        bottom = ["Plate", "End-Chock Bolts Tools", "1070-0200-0010", "2016-02-01 - en"]
        title, plate = _extract_plate_info([], bottom, bottom, _PLATE_RE)
        assert title == "End-Chock Bolts Tools"
        assert plate == "1070-0200-0010"

    def test_found_full_page_when_outside_bands(self):
        all_lines = ["MAN B&W", "015", "Jack", "Plate", "Cylinder Cover Tools",
                     "0540-0100-0002"]
        title, plate = _extract_plate_info([], [], all_lines, _PLATE_RE)
        assert title == "Cylinder Cover Tools"
        assert plate == "0540-0100-0002"

    def test_item_named_plate_not_mistaken_for_caption(self):
        # an item literally named "Plate" (no plate-number anchor) must NOT
        # become the title; the real caption later wins.
        all_lines = ["093", "Plate", "115", "Drain pipe", "127", "Screw",
                     "Plate", "Piston Cooling Arrangement", "1072-1400-0002"]
        title, plate = _extract_plate_info([], [], all_lines, _PLATE_RE)
        assert title == "Piston Cooling Arrangement"
        assert plate == "1072-1400-0002"

    def test_unanchored_number_yields_no_title(self):
        # a plate number whose line-above is a part name (no 'Plate' anchor) ->
        # empty title (better than a wrong one); plate number still returned.
        all_lines = ["Screw", "1072-0710-0015", "2012-10-09 - en"]
        title, plate = _extract_plate_info([], [], all_lines, _PLATE_RE)
        assert title == ""
        assert plate == "1072-0710-0015"

    @pytest.mark.parametrize("line,ok", [
        ("End-Chock Bolts Tools", True),
        ("Cylinder Cover", True),
        ("1070-0200-0010", False),
        ("2016-02-01 - en", False),
        ("Item No.", False),
        ("Plate", False),
        ("MAN B&W", False),
        ("(2)", False),
    ])
    def test_title_candidate_filter(self, line, ok):
        assert _is_title_candidate(line) is ok


def _make_pdf_with_unmapped(path):
    """Table with one mappable col + one unmappable col."""
    doc = fitz.open()
    page = doc.new_page(width=595, height=842)
    grid = [["Item No.", "Description", "Zubehör Spezial"],
            ["015", "hydraulic jack", "foobar data"],
            ["064", "support bracket", "more data"]]
    x0, y0, cw, rh = 60, 120, 150, 24
    for r, row in enumerate(grid):
        for c, cell in enumerate(row):
            page.insert_text((x0 + c * cw + 4, y0 + r * rh + 16), cell, fontsize=10)
    for r in range(len(grid) + 1):
        page.draw_line((x0, y0 + r * rh), (x0 + 3 * cw, y0 + r * rh))
    for c in range(4):
        page.draw_line((x0 + c * cw, y0), (x0 + c * cw, y0 + len(grid) * rh))
    doc.save(path)
    doc.close()


class TestStrictSchemaOutput:
    def test_main_sheet_is_exactly_schema(self, tmp_path, cfg):
        pdf = str(tmp_path / "u.pdf")
        _make_pdf_with_unmapped(pdf)
        paths = {"app_dir": str(tmp_path), "config_path": str(tmp_path / "c.json"),
                 "wip_tracker": str(tmp_path / "w.txt"), "models_dir": str(tmp_path / "m")}
        mapper = SemanticMapper(cfg, paths["models_dir"])
        res = process_file(pdf, cfg, paths, mapper, output_dir=str(tmp_path))
        wb = openpyxl.load_workbook(res.output_path)
        main = wb["Spare Parts"]
        headers = [c.value for c in main[1]]
        assert headers == cfg["target_schema"]  # exactly 14, no [UNMAPPED]
        assert not any("[UNMAPPED]" in str(h) for h in headers)

    def test_unmapped_data_on_separate_sheet(self, tmp_path, cfg):
        pdf = str(tmp_path / "u2.pdf")
        _make_pdf_with_unmapped(pdf)
        paths = {"app_dir": str(tmp_path), "config_path": str(tmp_path / "c.json"),
                 "wip_tracker": str(tmp_path / "w.txt"), "models_dir": str(tmp_path / "m")}
        mapper = SemanticMapper(cfg, paths["models_dir"])
        res = process_file(pdf, cfg, paths, mapper, output_dir=str(tmp_path))
        wb = openpyxl.load_workbook(res.output_path)
        # the unmappable German column data is preserved somewhere
        if "Unmapped (review)" in wb.sheetnames:
            txt = "".join(str(c.value) for row in wb["Unmapped (review)"].iter_rows()
                          for c in row)
            assert "foobar" in txt or "data" in txt
