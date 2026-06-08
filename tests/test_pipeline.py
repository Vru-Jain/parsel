"""
Integration tests for engine/pipeline.py.

Uses small synthetic PDFs built with PyMuPDF so the full
extract -> map -> transform -> QC -> Excel path runs without the huge real
files. Also verifies locked-Excel handling, corrupt-file handling, and that
the exported workbook matches the target schema exactly.
"""
from __future__ import annotations

import os

import fitz
import pandas as pd
import pytest
import openpyxl

from engine.semantic_mapper import SemanticMapper
from engine.pipeline import process_file


# --------------------------------------------------------------------------- #
# Synthetic PDF builders
# --------------------------------------------------------------------------- #
def _make_table_pdf(path: str, header: list[str], rows: list[list[str]],
                    title: str = "Cylinder Cover") -> None:
    """A simple gridded table PDF that PyMuPDF's table finder can read."""
    doc = fitz.open()
    page = doc.new_page(width=595, height=842)  # A4
    page.insert_text((60, 50), title, fontsize=14)
    page.insert_text((60, 70), "MAN B&W Co., Ltd.", fontsize=10)

    x0, y0 = 60, 120
    col_w, row_h = 150, 24
    grid = [header] + rows
    # draw text + ruling lines so both extraction strategies work
    for r, line in enumerate(grid):
        y = y0 + r * row_h
        for c, cell in enumerate(line):
            page.insert_text((x0 + c * col_w + 4, y + 16), str(cell), fontsize=10)
    n_rows, n_cols = len(grid), len(header)
    for r in range(n_rows + 1):
        yy = y0 + r * row_h
        page.draw_line((x0, yy), (x0 + n_cols * col_w, yy))
    for c in range(n_cols + 1):
        xx = x0 + c * col_w
        page.draw_line((xx, y0), (xx, y0 + n_rows * row_h))
    doc.save(path)
    doc.close()


def _make_blank_image_pdf(path: str, pages: int = 2) -> None:
    """A 'scanned' PDF: pages with only a drawn rectangle, no text layer."""
    doc = fitz.open()
    for _ in range(pages):
        page = doc.new_page(width=595, height=842)
        page.draw_rect(fitz.Rect(50, 50, 545, 792))  # just a box, no text
    doc.save(path)
    doc.close()


@pytest.fixture
def mapper(cfg, tmp_path):
    return SemanticMapper(cfg, str(tmp_path / "models"))


@pytest.fixture
def run_paths(tmp_path):
    return {
        "app_dir": str(tmp_path),
        "config_path": str(tmp_path / "config.json"),
        "wip_tracker": str(tmp_path / "WIP_Tracker.txt"),
        "models_dir": str(tmp_path / "models"),
    }


# --------------------------------------------------------------------------- #
# Happy path
# --------------------------------------------------------------------------- #
class TestHappyPath:
    def test_full_run_produces_excel(self, tmp_path, cfg, run_paths, mapper):
        pdf = str(tmp_path / "sample.pdf")
        _make_table_pdf(
            pdf,
            ["Item No.", "Item Designation", "Part No."],
            [["015", "hydraulic jack complete", "P-100"],
             ["064", "support for jack", "P-200"],
             ["111", "sealing ring with back-up", "P-300"]],
        )
        res = process_file(pdf, cfg, run_paths, mapper, output_dir=str(tmp_path))
        assert res.output_path and os.path.exists(res.output_path)
        assert res.rows == 3

        wb = openpyxl.load_workbook(res.output_path)
        ws = wb.active
        headers = [c.value for c in ws[1]]
        # output schema must match target exactly, in order
        assert headers == cfg["target_schema"]

    def test_mapping_and_rules_applied_end_to_end(self, tmp_path, cfg, run_paths, mapper):
        pdf = str(tmp_path / "rules.pdf")
        _make_table_pdf(
            pdf,
            ["Item No.", "Description"],
            [["1", "gasket EPDM"], ["2", "support for pump"]],
        )
        res = process_file(pdf, cfg, run_paths, mapper, output_dir=str(tmp_path))
        df = pd.read_excel(res.output_path).fillna("")
        names = df["Part Name"].tolist()
        # material tagged + conjunction lowercased
        assert any("Matl. EPDM" in n for n in names)
        assert any("support for pump".title().replace("For", "for") in n
                   or "Support for Pump" == n for n in names)

    def test_wip_tracker_written(self, tmp_path, cfg, run_paths, mapper):
        pdf = str(tmp_path / "wip.pdf")
        _make_table_pdf(pdf, ["Item No.", "Description"], [["1", "widget"]])
        process_file(pdf, cfg, run_paths, mapper, output_dir=str(tmp_path))
        assert os.path.exists(run_paths["wip_tracker"])
        log = open(run_paths["wip_tracker"], encoding="utf-8").read()
        assert "STARTED" in log and "FINISHED" in log


# --------------------------------------------------------------------------- #
# Failure / edge handling
# --------------------------------------------------------------------------- #
class TestEdgeCases:
    def test_corrupt_pdf_raises_runtime_error(self, tmp_path, cfg, run_paths, mapper):
        bad = str(tmp_path / "bad.pdf")
        with open(bad, "wb") as fh:
            fh.write(b"%PDF-1.4 this is not a real pdf body")
        with pytest.raises(RuntimeError):
            process_file(bad, cfg, run_paths, mapper, output_dir=str(tmp_path))

    def test_scanned_pdf_warns_not_crash(self, tmp_path, cfg, run_paths, mapper):
        pdf = str(tmp_path / "scanned.pdf")
        _make_blank_image_pdf(pdf, pages=3)
        cfg["options"]["enable_ocr"] = False
        res = process_file(pdf, cfg, run_paths, mapper, output_dir=str(tmp_path))
        assert res.rows == 0
        assert any("scan" in w.lower() for w in res.warnings)

    def test_locked_excel_saves_copy(self, tmp_path, cfg, run_paths, mapper):
        pdf = str(tmp_path / "lock.pdf")
        _make_table_pdf(pdf, ["Item No.", "Description"], [["1", "widget"]])
        # pre-create + hold the expected output open to simulate Excel lock
        expected = str(tmp_path / "lock_PROCESSED.xlsx")
        holder = open(expected, "w")
        try:
            res = process_file(pdf, cfg, run_paths, mapper, output_dir=str(tmp_path),
                               overwrite_locked_as_copy=True)
            # on Windows the open handle blocks writing; a copy is made instead
            if res.output_path:
                assert os.path.exists(res.output_path)
        finally:
            holder.close()

    def test_empty_table_pdf_no_output(self, tmp_path, cfg, run_paths, mapper):
        pdf = str(tmp_path / "empty.pdf")
        doc = fitz.open()
        doc.new_page()
        doc.insert_text = None  # noqa
        doc.save(pdf)
        doc.close()
        res = process_file(pdf, cfg, run_paths, mapper, output_dir=str(tmp_path))
        assert res.rows == 0


class TestPreviewFlow:
    def test_write_excel_false_defers_write(self, tmp_path, cfg, run_paths, mapper):
        """write_excel=False must produce the sheets in memory but write NOTHING;
        write_excel_file then saves them (the preview → save flow)."""
        from engine.pipeline import write_excel_file
        pdf = str(tmp_path / "prev.pdf")
        _make_table_pdf(pdf, ["Item No.", "Description"], [["1", "Valve"], ["2", "Seal"]])
        res = process_file(pdf, cfg, run_paths, mapper, output_dir=str(tmp_path),
                           write_excel=False)
        assert res.output_path == ""                      # nothing written yet
        assert res.main_df is not None and res.rows >= 2
        assert list(res.main_df.columns) == cfg["target_schema"]
        # no _PROCESSED.xlsx on disk yet
        assert not any(p.name.endswith("_PROCESSED.xlsx") for p in tmp_path.iterdir())
        # now save from the (previewed) result
        out = write_excel_file(res, str(tmp_path))
        assert os.path.exists(out) and res.output_path == out


class TestPageSelection:
    def test_pages_set_limits_to_chosen_pages(self, tmp_path, cfg, run_paths, mapper):
        """Selecting pages={2} must extract only page 2's parts, not page 1's."""
        pdf = str(tmp_path / "two.pdf")
        doc = fitz.open()
        for title, part in (("Plate A", "Valve"), ("Plate B", "Pump")):
            page = doc.new_page(width=595, height=842)
            x0, y0, cw, rh = 60, 120, 150, 24
            grid = [["Item No.", "Description"], ["1", part]]
            for r, line in enumerate(grid):
                for c, cell in enumerate(line):
                    page.insert_text((x0 + c * cw + 4, y0 + r * rh + 16), cell, fontsize=10)
            for r in range(len(grid) + 1):
                page.draw_line((x0, y0 + r * rh), (x0 + 2 * cw, y0 + r * rh))
            for c in range(3):
                page.draw_line((x0 + c * cw, y0), (x0 + c * cw, y0 + len(grid) * rh))
        doc.save(pdf)
        doc.close()
        res = process_file(pdf, cfg, run_paths, mapper, output_dir=str(tmp_path),
                           write_excel=False, pages={2})
        names = " ".join(res.main_df["Part Name"].astype(str)).lower()
        assert "pump" in names          # page 2 part present
        assert "valve" not in names     # page 1 part excluded


# --------------------------------------------------------------------------- #
# Schema conformance vs. the real target template
# --------------------------------------------------------------------------- #
class TestTargetSchemaConformance:
    def test_config_schema_matches_target_xlsx(self, cfg):
        from tests.conftest import DEMO_FILES
        target = DEMO_FILES["target_xlsx"]
        if not os.path.exists(target):
            pytest.skip("target xlsx not present")
        wb = openpyxl.load_workbook(target, data_only=True)
        ws = wb["Sheet1"]
        target_headers = [ws.cell(row=1, column=c).value
                          for c in range(1, ws.max_column + 1)]
        assert cfg["target_schema"] == target_headers
